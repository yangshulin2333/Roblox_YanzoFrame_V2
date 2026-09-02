from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import tempfile
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


IDENTIFIER_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
SUPPORTED_TYPES = {"string", "number", "boolean", "string[]", "number[]"}
ARRAY_ELEMENT_TYPES = {"string[]": "string", "number[]": "number"}
SUPPORTED_SCOPES = {"Shared", "Server"}


HEADER_STYLES = {
    1: {"fill": "FFDDEBF7", "font": "FF17365D", "size": 11},
    2: {"fill": "FF1F4E78", "font": "FFFFFFFF", "size": 10},
    3: {"fill": "FFD9E2F3", "font": "FF17365D", "size": 10},
}


@dataclass(frozen=True)
class ConfigIssue:
    code: str
    message: str
    sheet: str | None = None
    row: int | None = None
    column: int | None = None

    def format(self) -> str:
        parts = [self.code]
        if self.sheet is not None:
            parts.append(f"Sheet={self.sheet}")
        if self.row is not None and self.column is not None:
            parts.append(f"Cell={get_column_letter(self.column)}{self.row}")
        parts.append(self.message)
        return " | ".join(parts)


class ConfigFailure(Exception):
    def __init__(self, issues: list[ConfigIssue]):
        super().__init__("Config validation failed")
        self.issues = issues


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def load_schema(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ConfigFailure(
            [ConfigIssue("CONFIG_SCHEMA_INVALID", f"无法读取 Schema：{error}")]
        ) from error


def validate_schema(schema: dict[str, Any]) -> None:
    issues: list[ConfigIssue] = []
    if not isinstance(schema, dict) or schema.get("version") != 1:
        issues.append(ConfigIssue("CONFIG_SCHEMA_INVALID", "version 必须是 1"))
        raise ConfigFailure(issues)

    layout = schema.get("layout")
    expected_layout = {
        "descriptionRow": 1,
        "keyRow": 2,
        "typeRow": 3,
        "dataStartRow": 4,
    }
    if layout != expected_layout:
        issues.append(
            ConfigIssue(
                "CONFIG_SCHEMA_INVALID",
                "layout 必须固定为中文说明第 1 行、英文键名第 2 行、类型第 3 行、数据第 4 行起",
            )
        )

    sheets = schema.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        issues.append(ConfigIssue("CONFIG_SCHEMA_INVALID", "sheets 必须是非空数组"))
        raise ConfigFailure(issues)

    sheet_names: set[str] = set()
    sheet_map: dict[str, dict[str, Any]] = {}
    for sheet in sheets:
        if not isinstance(sheet, dict):
            issues.append(ConfigIssue("CONFIG_SCHEMA_INVALID", "每个 Sheet 定义必须是对象"))
            continue

        name = sheet.get("name")
        if not isinstance(name, str) or not IDENTIFIER_PATTERN.fullmatch(name):
            issues.append(
                ConfigIssue("CONFIG_SCHEMA_INVALID", f"非法 Sheet 名：{name!r}")
            )
            continue
        if name in sheet_names:
            issues.append(ConfigIssue("CONFIG_SCHEMA_INVALID", f"重复 Sheet：{name}"))
            continue
        sheet_names.add(name)
        sheet_map[name] = sheet

        workbook_name = sheet.get("workbook")
        if (
            not isinstance(workbook_name, str)
            or Path(workbook_name).name != workbook_name
            or Path(workbook_name).suffix.lower() != ".xlsx"
            or workbook_name.startswith("~$")
        ):
            issues.append(
                ConfigIssue(
                    "CONFIG_SCHEMA_INVALID",
                    "workbook 必须是 workbooks 目录下的 .xlsx 文件名",
                    sheet=name,
                )
            )

        if not isinstance(sheet.get("descriptionZh"), str) or not sheet[
            "descriptionZh"
        ].strip():
            issues.append(
                ConfigIssue("CONFIG_SCHEMA_INVALID", "Sheet 缺少中文说明", sheet=name)
            )
        if sheet.get("scope") not in SUPPORTED_SCOPES:
            issues.append(
                ConfigIssue(
                    "CONFIG_SCHEMA_INVALID",
                    "scope 只能是 Shared 或 Server",
                    sheet=name,
                )
            )

        columns = sheet.get("columns")
        if not isinstance(columns, list) or not columns:
            issues.append(
                ConfigIssue("CONFIG_SCHEMA_INVALID", "columns 必须是非空数组", sheet=name)
            )
            continue

        column_keys: set[str] = set()
        for column in columns:
            if not isinstance(column, dict):
                issues.append(
                    ConfigIssue("CONFIG_SCHEMA_INVALID", "字段定义必须是对象", sheet=name)
                )
                continue
            key = column.get("key")
            if not isinstance(key, str) or not IDENTIFIER_PATTERN.fullmatch(key):
                issues.append(
                    ConfigIssue("CONFIG_SCHEMA_INVALID", f"非法字段名：{key!r}", sheet=name)
                )
                continue
            if key in column_keys:
                issues.append(
                    ConfigIssue("CONFIG_SCHEMA_INVALID", f"重复字段：{key}", sheet=name)
                )
                continue
            column_keys.add(key)

            if not isinstance(column.get("descriptionZh"), str) or not column[
                "descriptionZh"
            ].strip():
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {key} 缺少中文释义",
                        sheet=name,
                    )
                )
            if column.get("type") not in SUPPORTED_TYPES:
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {key} 使用了不支持的类型",
                        sheet=name,
                    )
                )
            if column.get("unique") is True and column.get("type") in ARRAY_ELEMENT_TYPES:
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {key} 的数组类型不能声明 unique",
                        sheet=name,
                    )
                )
            if "enum" in column and (
                not isinstance(column["enum"], list) or not column["enum"]
            ):
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {key} 的 enum 必须是数组",
                        sheet=name,
                    )
                )
            if "enum" in column and column.get("type") in ARRAY_ELEMENT_TYPES:
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {key} 的数组类型暂不支持 enum",
                        sheet=name,
                    )
                )
            for limit_name in ("min", "max"):
                if limit_name in column and (
                    column.get("type") != "number"
                    or not isinstance(column[limit_name], (int, float))
                    or isinstance(column[limit_name], bool)
                ):
                    issues.append(
                        ConfigIssue(
                            "CONFIG_SCHEMA_INVALID",
                            f"字段 {key} 的 {limit_name} 只能用于 number",
                            sheet=name,
                        )
                    )

        primary_key = sheet.get("primaryKey")
        if primary_key not in column_keys:
            issues.append(
                ConfigIssue(
                    "CONFIG_SCHEMA_INVALID",
                    "primaryKey 必须指向本 Sheet 的字段",
                    sheet=name,
                )
            )
        else:
            primary_column = next(column for column in columns if column.get("key") == primary_key)
            if (
                primary_column.get("type") != "string"
                or primary_column.get("required") is not True
                or primary_column.get("unique") is not True
            ):
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        "primaryKey 必须是 required、unique 的 string 字段",
                        sheet=name,
                    )
                )

    # 引用只允许指向已声明且唯一的字段，避免生成后出现不稳定关联。
    for sheet_name, sheet in sheet_map.items():
        for column in sheet.get("columns", []):
            if not isinstance(column, dict):
                continue
            reference = column.get("ref")
            if reference is None:
                continue
            if not isinstance(reference, str) or reference.count(".") != 1:
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {column.get('key')} 的 ref 格式应为 Sheet.Field",
                        sheet=sheet_name,
                    )
                )
                continue
            target_sheet_name, target_key = reference.split(".")
            target_sheet = sheet_map.get(target_sheet_name)
            target_columns = target_sheet.get("columns", []) if target_sheet else []
            target_column = next(
                (
                    item
                    for item in target_columns
                    if isinstance(item, dict) and item.get("key") == target_key
                ),
                None,
            )
            if target_column is None or target_column.get("unique") is not True:
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {column.get('key')} 的 ref 必须指向 unique 字段",
                        sheet=sheet_name,
                    )
                )
            elif target_column.get("type") != column.get("type"):
                issues.append(
                    ConfigIssue(
                        "CONFIG_SCHEMA_INVALID",
                        f"字段 {column.get('key')} 与引用目标的类型必须一致",
                        sheet=sheet_name,
                    )
                )

    if issues:
        raise ConfigFailure(issues)


def style_header_cell(cell: Any, row_number: int) -> None:
    style = HEADER_STYLES[row_number]
    thin = Side(style="thin", color="FFB4C6E7")
    cell.font = Font(
        name="Microsoft YaHei",
        size=style["size"],
        bold=True,
        color=style["font"],
    )
    cell.fill = PatternFill("solid", fgColor=style["fill"])
    cell.alignment = Alignment(
        horizontal="left" if row_number == 1 else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_sheet_headers(
    worksheet: Any,
    sheet: dict[str, Any],
    layout: dict[str, int],
    start_column: int = 1,
) -> None:
    columns = sheet["columns"]
    for column_number in range(start_column, len(columns) + 1):
        column = columns[column_number - 1]
        values = {
            layout["descriptionRow"]: column["descriptionZh"],
            layout["keyRow"]: column["key"],
            layout["typeRow"]: column["type"],
        }
        for row_number, value in values.items():
            cell = worksheet.cell(row_number, column_number, value)
            style_header_cell(cell, row_number)

        description_width = len(column["descriptionZh"]) * 2 + 4
        worksheet.column_dimensions[get_column_letter(column_number)].width = max(
            14, min(42, description_width)
        )

    worksheet.row_dimensions[layout["descriptionRow"]].height = 34
    worksheet.row_dimensions[layout["keyRow"]].height = 24
    worksheet.row_dimensions[layout["typeRow"]].height = 22


def non_blank_header_keys(worksheet: Any, key_row: int) -> list[Any]:
    keys = [
        clean_value(worksheet.cell(key_row, column).value)
        for column in range(1, worksheet.max_column + 1)
    ]
    while keys and is_blank(keys[-1]):
        keys.pop()
    return keys


def save_workbook_atomic(workbook: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.config-sync-{uuid.uuid4().hex}.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    except OSError as error:
        raise ConfigFailure(
            [
                ConfigIssue(
                    "CONFIG_WORKBOOK_WRITE_FAILED",
                    f"无法更新 {path.name}，请关闭 Excel 后重试：{error}",
                )
            ]
        ) from error
    finally:
        temporary.unlink(missing_ok=True)


def sync_workbooks_from_schema(
    directory: Path, schema: dict[str, Any]
) -> list[str]:
    """只补齐安全缺失的结构；绝不移动、删除或覆盖已有数据。"""
    directory.mkdir(parents=True, exist_ok=True)
    layout = schema["layout"]
    by_workbook: dict[str, list[dict[str, Any]]] = {}
    for sheet in schema["sheets"]:
        by_workbook.setdefault(sheet["workbook"], []).append(sheet)

    changes: list[str] = []
    for workbook_name, sheets in by_workbook.items():
        path = directory / workbook_name
        if path.exists():
            try:
                workbook = load_workbook(path, data_only=False, read_only=False)
            except Exception as error:
                raise ConfigFailure(
                    [
                        ConfigIssue(
                            "CONFIG_WORKBOOK_READ_FAILED",
                            f"无法读取 Excel：{error}",
                        )
                    ]
                ) from error
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)

        workbook_changed = False
        try:
            for sheet in sheets:
                sheet_name = sheet["name"]
                if sheet_name not in workbook.sheetnames:
                    worksheet = workbook.create_sheet(sheet_name)
                    write_sheet_headers(worksheet, sheet, layout)
                    worksheet.freeze_panes = f"A{layout['dataStartRow']}"
                    changes.append(
                        f"CONFIG_WORKBOOK_SYNCED | Workbook={workbook_name} | "
                        f"Sheet={sheet_name} | Action=CREATE_SHEET"
                    )
                    workbook_changed = True
                    continue

                worksheet = workbook[sheet_name]
                actual_keys = non_blank_header_keys(worksheet, layout["keyRow"])
                expected_keys = [column["key"] for column in sheet["columns"]]

                if not actual_keys:
                    write_sheet_headers(worksheet, sheet, layout)
                    worksheet.freeze_panes = f"A{layout['dataStartRow']}"
                    changes.append(
                        f"CONFIG_WORKBOOK_SYNCED | Workbook={workbook_name} | "
                        f"Sheet={sheet_name} | Action=CREATE_HEADERS"
                    )
                    workbook_changed = True
                elif (
                    len(actual_keys) < len(expected_keys)
                    and expected_keys[: len(actual_keys)] == actual_keys
                ):
                    write_sheet_headers(
                        worksheet,
                        sheet,
                        layout,
                        start_column=len(actual_keys) + 1,
                    )
                    changes.append(
                        f"CONFIG_WORKBOOK_SYNCED | Workbook={workbook_name} | "
                        f"Sheet={sheet_name} | Action=APPEND_COLUMNS"
                    )
                    workbook_changed = True

            if workbook_changed:
                if workbook.worksheets:
                    workbook.active = 0
                save_workbook_atomic(workbook, path)
        finally:
            workbook.close()

    return changes


def read_workbook_model(path: Path, layout: dict[str, int]) -> dict[str, Any]:
    try:
        # 普通模式能稳定读取由不同工具生成的工作表尺寸；这里只读取，不保存 Excel。
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception as error:
        raise ConfigFailure(
            [ConfigIssue("CONFIG_WORKBOOK_READ_FAILED", f"无法读取 Excel：{error}")]
        ) from error

    model: dict[str, Any] = {"sheetNames": list(workbook.sheetnames), "sheets": {}}
    try:
        for worksheet in workbook.worksheets:
            formulas: set[tuple[int, int]] = set()
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formulas.add((cell.row, cell.column))

            model["sheets"][worksheet.title] = {
                "descriptions": [
                    worksheet.cell(layout["descriptionRow"], column).value
                    for column in range(1, worksheet.max_column + 1)
                ],
                "keys": [
                    worksheet.cell(layout["keyRow"], column).value
                    for column in range(1, worksheet.max_column + 1)
                ],
                "types": [
                    worksheet.cell(layout["typeRow"], column).value
                    for column in range(1, worksheet.max_column + 1)
                ],
                "rows": [
                    {
                        "row": row_number,
                        "values": [
                            worksheet.cell(row_number, column).value
                            for column in range(1, worksheet.max_column + 1)
                        ],
                    }
                    for row_number in range(
                        layout["dataStartRow"], worksheet.max_row + 1
                    )
                ],
                "formulas": formulas,
                "workbook": path.name,
            }
    finally:
        workbook.close()

    return model


def read_workbooks_model(directory: Path, layout: dict[str, int]) -> dict[str, Any]:
    if not directory.is_dir():
        raise ConfigFailure(
            [ConfigIssue("CONFIG_WORKBOOK_DIR_INVALID", f"找不到 Excel 目录：{directory}")]
        )

    workbook_paths = sorted(
        path
        for path in directory.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and not path.name.startswith(".")
    )
    if not workbook_paths:
        raise ConfigFailure(
            [ConfigIssue("CONFIG_WORKBOOK_MISSING", "Excel 目录中没有可用的 .xlsx 文件")]
        )

    # 所有工作簿共用一个 Sheet 命名空间，避免生成同名 Luau 时互相覆盖。
    model: dict[str, Any] = {"sheetNames": [], "sheets": {}}
    issues: list[ConfigIssue] = []
    for workbook_path in workbook_paths:
        workbook_model = read_workbook_model(workbook_path, layout)
        for sheet_name in workbook_model["sheetNames"]:
            if sheet_name in model["sheets"]:
                previous = model["sheets"][sheet_name]["workbook"]
                issues.append(
                    ConfigIssue(
                        "CONFIG_DUPLICATE_SHEET",
                        f"Sheet 同时存在于 {previous} 和 {workbook_path.name}",
                        sheet=sheet_name,
                    )
                )
                continue
            model["sheetNames"].append(sheet_name)
            model["sheets"][sheet_name] = workbook_model["sheets"][sheet_name]

    if issues:
        raise ConfigFailure(issues)
    return model


def parse_array_value(value: Any, expected_type: str) -> list[Any] | None:
    if not isinstance(value, str):
        return None

    parts = [part.strip() for part in value.split(",")]
    if not parts or any(part == "" for part in parts):
        return None
    if expected_type == "string[]":
        return parts

    numbers: list[int | float] = []
    for part in parts:
        try:
            number = float(part)
        except ValueError:
            return None
        if not math.isfinite(number):
            return None
        numbers.append(int(number) if number.is_integer() else number)
    return numbers


def validate_value(
    raw_value: Any,
    column: dict[str, Any],
    sheet_name: str,
    row_number: int,
    column_number: int,
) -> tuple[Any, list[ConfigIssue]]:
    issues: list[ConfigIssue] = []
    key = column["key"]
    if is_blank(raw_value):
        if column.get("required") is True:
            issues.append(
                ConfigIssue(
                    "CONFIG_REQUIRED_MISSING",
                    f"字段 {key} 为必填项",
                    sheet_name,
                    row_number,
                    column_number,
                )
            )
        return None, issues

    expected_type = column["type"]
    value = clean_value(raw_value)
    if expected_type == "boolean" and isinstance(value, str):
        if value.upper() == "TRUE":
            value = True
        elif value.upper() == "FALSE":
            value = False
    elif expected_type in ARRAY_ELEMENT_TYPES:
        value = parse_array_value(value, expected_type)
    valid_type = (
        (expected_type == "string" and isinstance(value, str))
        or (
            expected_type == "number"
            and isinstance(value, (int, float))
            and not isinstance(value, bool)
            and math.isfinite(value)
        )
        or (expected_type == "boolean" and isinstance(value, bool))
        or (expected_type in ARRAY_ELEMENT_TYPES and isinstance(value, list))
    )
    if not valid_type:
        issues.append(
            ConfigIssue(
                "CONFIG_TYPE_INVALID",
                f"字段 {key} 必须是 {expected_type}",
                sheet_name,
                row_number,
                column_number,
            )
        )
        return None, issues

    if "min" in column and value < column["min"]:
        issues.append(
            ConfigIssue(
                "CONFIG_NUMBER_OUT_OF_RANGE",
                f"字段 {key} 不能小于 {column['min']}",
                sheet_name,
                row_number,
                column_number,
            )
        )
    if "max" in column and value > column["max"]:
        issues.append(
            ConfigIssue(
                "CONFIG_NUMBER_OUT_OF_RANGE",
                f"字段 {key} 不能大于 {column['max']}",
                sheet_name,
                row_number,
                column_number,
            )
        )
    if "enum" in column and value not in column["enum"]:
        allowed = ", ".join(str(item) for item in column["enum"])
        issues.append(
            ConfigIssue(
                "CONFIG_ENUM_INVALID",
                f"字段 {key} 只能填写：{allowed}",
                sheet_name,
                row_number,
                column_number,
            )
        )

    return value, issues


def validate_workbook_model(
    model: dict[str, Any], schema: dict[str, Any]
) -> tuple[dict[str, list[dict[str, Any]]], list[ConfigIssue]]:
    issues: list[ConfigIssue] = []
    typed_rows: dict[str, list[dict[str, Any]]] = {}
    expected_names = [sheet["name"] for sheet in schema["sheets"]]
    actual_names = model.get("sheetNames", [])

    for name in actual_names:
        if name not in expected_names:
            issues.append(ConfigIssue("CONFIG_UNKNOWN_SHEET", "Excel 中存在未知 Sheet", name))
    for name in expected_names:
        if name not in actual_names:
            issues.append(ConfigIssue("CONFIG_SHEET_MISSING", "Excel 缺少必要 Sheet", name))

    for sheet in schema["sheets"]:
        sheet_name = sheet["name"]
        actual = model.get("sheets", {}).get(sheet_name)
        if actual is None:
            continue

        columns = sheet["columns"]
        expected_descriptions = [column["descriptionZh"] for column in columns]
        expected_keys = [column["key"] for column in columns]
        expected_types = [column["type"] for column in columns]
        actual_descriptions = actual.get("descriptions", [])
        actual_keys = actual.get("keys", [])
        actual_types = actual.get("types", [])

        if actual.get("workbook") != sheet["workbook"]:
            issues.append(
                ConfigIssue(
                    "CONFIG_SHEET_WRONG_WORKBOOK",
                    f"Sheet 应放在 {sheet['workbook']}，当前位于 {actual.get('workbook')}",
                    sheet_name,
                )
            )

        for column_number, expected in enumerate(expected_descriptions, start=1):
            found = actual_descriptions[column_number - 1] if column_number <= len(actual_descriptions) else None
            if clean_value(found) != expected:
                issues.append(
                    ConfigIssue(
                        "CONFIG_DESCRIPTION_MISMATCH",
                        f"中文释义应为：{expected}",
                        sheet_name,
                        1,
                        column_number,
                    )
                )
        for column_number, expected in enumerate(expected_keys, start=1):
            found = actual_keys[column_number - 1] if column_number <= len(actual_keys) else None
            if clean_value(found) != expected:
                issues.append(
                    ConfigIssue(
                        "CONFIG_HEADER_MISMATCH",
                        f"英文键名应为：{expected}",
                        sheet_name,
                        2,
                        column_number,
                    )
                )
        for column_number, expected in enumerate(expected_types, start=1):
            found = actual_types[column_number - 1] if column_number <= len(actual_types) else None
            if clean_value(found) != expected:
                issues.append(
                    ConfigIssue(
                        "CONFIG_TYPE_HEADER_MISMATCH",
                        f"字段类型应为：{expected}",
                        sheet_name,
                        3,
                        column_number,
                    )
                )

        max_actual_columns = max(
            [len(actual_descriptions), len(actual_keys), len(actual_types)]
            + [len(row.get("values", [])) for row in actual.get("rows", [])]
        )
        for column_number in range(len(columns) + 1, max_actual_columns + 1):
            description = actual_descriptions[column_number - 1] if column_number <= len(actual_descriptions) else None
            key = actual_keys[column_number - 1] if column_number <= len(actual_keys) else None
            field_type = actual_types[column_number - 1] if column_number <= len(actual_types) else None
            data_row = next(
                (
                    row
                    for row in actual.get("rows", [])
                    if column_number <= len(row.get("values", []))
                    and not is_blank(row["values"][column_number - 1])
                ),
                None,
            )
            if (
                not is_blank(description)
                or not is_blank(key)
                or not is_blank(field_type)
                or data_row is not None
            ):
                issue_row = 1
                if is_blank(description) and not is_blank(key):
                    issue_row = 2
                elif is_blank(description) and is_blank(key) and not is_blank(field_type):
                    issue_row = 3
                elif is_blank(description) and is_blank(key) and is_blank(field_type):
                    issue_row = data_row["row"] if data_row is not None else 3
                issues.append(
                    ConfigIssue(
                        "CONFIG_UNKNOWN_COLUMN",
                        "Excel 中存在 Schema 未声明的字段",
                        sheet_name,
                        issue_row,
                        column_number,
                    )
                )

        formula_cells = actual.get("formulas", set())
        for row_number, column_number in sorted(formula_cells):
            issues.append(
                ConfigIssue(
                    "CONFIG_FORMULA_NOT_ALLOWED",
                    "Config 不允许使用 Excel 公式",
                    sheet_name,
                    row_number,
                    column_number,
                )
            )

        unique_values: dict[str, dict[Any, int]] = {
            column["key"]: {} for column in columns if column.get("unique") is True
        }
        sheet_rows: list[dict[str, Any]] = []
        for raw_row in actual.get("rows", []):
            row_number = raw_row["row"]
            values = raw_row.get("values", [])
            if all(is_blank(value) for value in values):
                continue

            typed: dict[str, Any] = {"__row": row_number}
            for column_number, column in enumerate(columns, start=1):
                if (row_number, column_number) in formula_cells:
                    continue
                raw_value = values[column_number - 1] if column_number <= len(values) else None
                value, value_issues = validate_value(
                    raw_value, column, sheet_name, row_number, column_number
                )
                issues.extend(value_issues)
                if value is not None:
                    typed[column["key"]] = value

                    if column.get("unique") is True:
                        seen = unique_values[column["key"]]
                        if value in seen:
                            issues.append(
                                ConfigIssue(
                                    "CONFIG_DUPLICATE_VALUE",
                                    f"字段 {column['key']} 与第 {seen[value]} 行重复",
                                    sheet_name,
                                    row_number,
                                    column_number,
                                )
                            )
                        else:
                            seen[value] = row_number
            sheet_rows.append(typed)
        typed_rows[sheet_name] = sheet_rows

    # 所有 Sheet 都完成类型校验后，再检查跨表引用。
    for sheet in schema["sheets"]:
        sheet_name = sheet["name"]
        for column_number, column in enumerate(sheet["columns"], start=1):
            reference = column.get("ref")
            if reference is None:
                continue
            target_sheet_name, target_key = reference.split(".")
            target_values = {
                row.get(target_key)
                for row in typed_rows.get(target_sheet_name, [])
                if row.get(target_key) is not None
            }
            for row in typed_rows.get(sheet_name, []):
                value = row.get(column["key"])
                if value is not None and value not in target_values:
                    issues.append(
                        ConfigIssue(
                            "CONFIG_REFERENCE_NOT_FOUND",
                            f"字段 {column['key']} 找不到引用 {reference}={value}",
                            sheet_name,
                            row["__row"],
                            column_number,
                        )
                    )

    return typed_rows, issues


def lua_value(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    if isinstance(value, str):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        return "{ " + ", ".join(lua_value(item) for item in value) + " }"
    raise TypeError(f"Unsupported Luau value: {value!r}")


def build_outputs(
    schema: dict[str, Any], typed_rows: dict[str, list[dict[str, Any]]]
) -> dict[str, dict[str, str]]:
    outputs: dict[str, dict[str, str]] = {"Shared": {}, "Server": {}}
    for sheet in schema["sheets"]:
        sheet_name = sheet["name"]
        primary_key = sheet["primaryKey"]
        lines = [
            "--!strict",
            "-- 此文件由 Excel Config 工具自动生成，请勿手动修改。",
            f"-- 来源：design/config/workbooks/{sheet['workbook']} / {sheet_name}",
            "",
        ]
        rows = typed_rows[sheet_name]
        if not rows:
            lines.extend(["return {}", ""])
        else:
            lines.append("return {")
            for row in rows:
                lines.append(f"\t[{lua_value(row[primary_key])}] = {{")
                for column in sheet["columns"]:
                    key = column["key"]
                    if key in row:
                        lines.append(f"\t\t{key} = {lua_value(row[key])},")
                lines.append("\t},")
            lines.extend(["}", ""])
        outputs[sheet["scope"]][f"{sheet_name}.lua"] = "\n".join(lines)
    return outputs


def output_targets(repo_root: Path) -> dict[str, Path]:
    return {
        "Shared": repo_root
        / "src/ReplicatedStorage/Game/Shared/Config/Generated",
        "Server": repo_root
        / "src/ServerScriptService/Server/Game/Config/Generated",
    }


def check_outputs(repo_root: Path, outputs: dict[str, dict[str, str]]) -> list[ConfigIssue]:
    issues: list[ConfigIssue] = []
    for scope, target in output_targets(repo_root).items():
        expected_files = outputs[scope]
        actual_files = {
            path.name for path in target.iterdir() if path.is_file()
        } if target.exists() else set()
        if actual_files != set(expected_files):
            issues.append(
                ConfigIssue(
                    "CONFIG_OUTPUT_OUTDATED",
                    f"{scope} 生成文件集合与 Excel 不一致",
                )
            )
            continue
        for file_name, expected in expected_files.items():
            if (target / file_name).read_text(encoding="utf-8") != expected:
                issues.append(
                    ConfigIssue(
                        "CONFIG_OUTPUT_OUTDATED",
                        f"{scope}/{file_name} 与 Excel 不一致",
                    )
                )
    return issues


def write_outputs_atomic(repo_root: Path, outputs: dict[str, dict[str, str]]) -> None:
    # 临时文件和旧目录都留在 src 外，避免 Rojo/Luau 监听到短暂备份路径。
    temp_root = Path(tempfile.mkdtemp(prefix=".config-build-", dir=repo_root))
    targets = output_targets(repo_root)
    replaced: list[tuple[Path, Path | None]] = []
    try:
        for scope, files in outputs.items():
            temp_scope = temp_root / scope
            temp_scope.mkdir(parents=True)
            for file_name, content in files.items():
                (temp_scope / file_name).write_text(content, encoding="utf-8", newline="\n")

        for scope, target in targets.items():
            target.parent.mkdir(parents=True, exist_ok=True)
            backup = temp_root / f"{scope}.previous"
            previous_backup: Path | None = None
            if target.exists():
                os.replace(target, backup)
                previous_backup = backup
            try:
                os.replace(temp_root / scope, target)
            except Exception:
                if previous_backup is not None and previous_backup.exists():
                    os.replace(previous_backup, target)
                raise
            replaced.append((target, previous_backup))

        # 两个目录都替换成功后即视为提交完成；旧备份清理失败不能再触发回滚。
        for _, backup in replaced:
            if backup is not None and backup.exists():
                try:
                    shutil.rmtree(backup)
                except OSError:
                    pass
    except Exception:
        for target, backup in reversed(replaced):
            if target.exists():
                shutil.rmtree(target)
            if backup is not None and backup.exists():
                os.replace(backup, target)
        raise
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def run(args: argparse.Namespace) -> int:
    repo_root = Path(args.repo_root).resolve()
    workbook_dir = Path(args.workbook_dir).resolve()
    schema_path = Path(args.schema).resolve()

    try:
        schema = load_schema(schema_path)
        validate_schema(schema)
        if not args.check:
            for change in sync_workbooks_from_schema(workbook_dir, schema):
                print(change)
        model = read_workbooks_model(workbook_dir, schema["layout"])
        typed_rows, issues = validate_workbook_model(model, schema)
        if issues:
            raise ConfigFailure(issues)
        outputs = build_outputs(schema, typed_rows)

        if args.check:
            check_issues = check_outputs(repo_root, outputs)
            if check_issues:
                raise ConfigFailure(check_issues)
            print("CONFIG_OUTPUT_CHECK_OK")
            return 0

        write_outputs_atomic(repo_root, outputs)
        print("CONFIG_IMPORT_OK")
        return 0
    except ConfigFailure as error:
        for issue in error.issues:
            print(issue.format())
        return 1
    except Exception as error:
        print(ConfigIssue("CONFIG_OUTPUT_FAILED", f"生成失败：{error}").format())
        return 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将三行表头 Excel Config 生成确定性的 Luau 模块")
    parser.add_argument("--repo-root", required=True)
    parser.add_argument("--workbook-dir", required=True)
    parser.add_argument("--schema", required=True)
    parser.add_argument("--check", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(run(parse_args()))
